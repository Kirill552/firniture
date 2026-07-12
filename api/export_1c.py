"""1C export core — Task 12.

Deterministic export of ManufacturingSpec to 1C:Enterprise compatible CSV.

1C:Enterprise uses semicolon-delimited CSV with Russian headers.
Columns follow standard 1C import format for «Номенклатура» (items).

Domain:
- Export1CRecord: stable-ID export record for 1C
- ExportStatus: draft/approved decision domain
- build_1c_records: extract items from spec
- export_1c_csv: produce Russian CSV (UTF-8 BOM, semicolon delimiter)
"""
from __future__ import annotations

import csv
import hashlib
import io
from datetime import UTC, datetime
from enum import Enum
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field, field_validator

from api.manufacturing.contracts import ManufacturingSpec

# ── Enums ─────────────────────────────────────────────────────────────


class ExportStatus(str, Enum):
    """Статус экспорта: черновик или утверждён."""

    DRAFT = "draft"
    APPROVED = "approved"


# ── Models ────────────────────────────────────────────────────────────


class Export1CRecord(BaseModel):
    """Запись экспорта для 1C:Enterprise со стабильным ID."""

    record_id: str
    panel_id: str
    name: str
    width_mm: float
    height_mm: float
    thickness_mm: float
    material: str
    quantity: int = Field(1, ge=1)
    status: ExportStatus = ExportStatus.DRAFT

    @field_validator("record_id", mode="before")
    @classmethod
    def _validate_record_id(cls, v: str) -> str:
        if not v or not v.strip():
            raise ValueError("record_id не может быть пустым")
        return v

    @field_validator("quantity", mode="before")
    @classmethod
    def _validate_quantity(cls, v: int) -> int:
        if v < 1:
            raise ValueError(f"quantity должно быть >= 1, получено {v}")
        return v


# ── Стабильные ID ─────────────────────────────────────────────────────


def make_1c_record_id(panel_id: str, revision: int = 1) -> str:
    """Детерминированный ID записи 1C (stable export identifier).

    Включает revision для идентичности в BOM/Excel/CSV/CAM manifest.
    CAM manifest здесь — только контракт идентификаторов (не реализация CAM/G-code).
    """
    raw = f"{panel_id}:r{revision}"
    h = hashlib.sha256(raw.encode()).hexdigest()[:12]
    return f"1c_{panel_id}_r{revision}_{h}"


# ── Построение записей ─────────────────────────────────────────────────


def build_1c_records(
    spec: ManufacturingSpec, *, revision: int = 1, status: ExportStatus = ExportStatus.DRAFT
) -> list[Export1CRecord]:
    """Построить записи для 1C из ManufacturingSpec.

    Детерминированный порядок: панели сортируются по ID.
    revision + status для stable IDs и production/stale gate.
    """
    records: list[Export1CRecord] = []
    for panel in sorted(spec.panels, key=lambda p: p.id):
        records.append(
            Export1CRecord(
                record_id=make_1c_record_id(panel.id, revision),
                panel_id=panel.id,
                name=f"Панель {panel.id}",
                width_mm=panel.width_mm,
                height_mm=panel.height_mm,
                thickness_mm=panel.thickness_mm,
                material=panel.material or "ЛДСП",
                status=status,
            )
        )
    return records


# ── CSV экспорт ───────────────────────────────────────────────────────

_CSV_HEADER = [
    "ИД записи",
    "ИД панели",
    "Наименование",
    "Длина, мм",
    "Ширина, мм",
    "Толщина, мм",
    "Материал",
    "Кол-во",
    "Статус",
]

_STATUS_RU: dict[ExportStatus, str] = {
    ExportStatus.DRAFT: "Черновик",
    ExportStatus.APPROVED: "Утверждено",
}


def export_1c_csv(records: list[Export1CRecord], *, decimal_sep: str = ".", watermark: bool = False) -> str:
    """Сформировать CSV строку для 1C:Enterprise (русская).

    Формат: UTF-8 BOM, разделитель — точка с запятой.
    decimal_sep поддерживает русские fixtures ('.' или ',').
    watermark добавляет метку preview draft.
    """
    def _fmt(v: float) -> str:
        return f"{v:.1f}".replace(".", decimal_sep)

    buf = io.StringIO()
    buf.write("\ufeff")
    if watermark:
        buf.write("ВОДЯНОЙ_ЗНАК;ПРЕДВАРИТЕЛЬНЫЙ ЧЕРНОВИК — НЕ ДЛЯ ПРОИЗВОДСТВА;;;;;;;\n")
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(_CSV_HEADER)
    for rec in records:
        writer.writerow(
            [
                rec.record_id,
                rec.panel_id,
                rec.name,
                _fmt(rec.width_mm),
                _fmt(rec.height_mm),
                _fmt(rec.thickness_mm),
                rec.material,
                str(rec.quantity),
                _STATUS_RU[rec.status],
            ]
        )
    return buf.getvalue()


def export_1c_csv_bytes(records: list[Export1CRecord], **kwargs: Any) -> bytes:
    """CSV как bytes (для записи в файл)."""
    return export_1c_csv(records, **kwargs).encode("utf-8")


# ── Order-level export API (used by routers.py) ─────────────────────


def generate_1c_filename(order: Any, ext: str) -> str:
    """Детерминированное имя файла для 1C-экспорта.

    Формат: 1C_Order_{customer_ref|short-id}_{YYYYMMDD}.{ext}
    """
    ref = getattr(order, "customer_ref", None) or str(order.id)[:8]
    # sanitise for filesystem
    ref = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in ref)
    ts = order.created_at.strftime("%Y%m%d") if hasattr(order, "created_at") and order.created_at else datetime.now(UTC).strftime("%Y%m%d")
    return f"1C_Order_{ref}_{ts}.{ext}"


# ── Headers ─────────────────────────────────────────────────────────

_XLSX_HEADER_FONT = Font(bold=True, color="FFFFFF")
_XLSX_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_XLSX_HEADER_ALIGN = Alignment(horizontal="center")


def _style_header(ws: Any, row: int = 1) -> None:
    """Apply header styling to the first row."""
    for cell in ws[row]:
        cell.font = _XLSX_HEADER_FONT
        cell.fill = _XLSX_HEADER_FILL
        cell.alignment = _XLSX_HEADER_ALIGN


def _write_order_sheet(wb: Workbook, order: Any) -> None:
    """Sheet «Заказ» — общая информация о заказе."""
    ws = wb.create_sheet("Заказ", 0)
    ws.append(["Поле", "Значение"])
    ws.append(["ID заказа", str(order.id)])
    ws.append(["Клиент", getattr(order, "customer_ref", "") or ""])
    ws.append(["Статус", getattr(order, "status", "")])
    ws.append(["Создан", str(order.created_at) if getattr(order, "created_at", None) else ""])
    ws.append(["Кол-во изделий", len(getattr(order, "products", []))])
    _style_header(ws)


def _write_products_sheet(wb: Workbook, products: list[Any]) -> None:
    """Sheet «Изделия» — конфигурация каждого изделия."""
    ws = wb.create_sheet("Изделия")
    ws.append(["ID", "Название", "Ширина, мм", "Высота, мм", "Глубина, мм", "Материал", "Толщина, мм", "Заметки"])
    for p in products:
        ws.append([
            str(p.id),
            getattr(p, "name", "") or "",
            getattr(p, "width_mm", 0),
            getattr(p, "height_mm", 0),
            getattr(p, "depth_mm", 0),
            getattr(p, "material", "") or "",
            getattr(p, "thickness_mm", None) or "",
            getattr(p, "notes", "") or "",
        ])
    _style_header(ws)


def _write_panels_sheet(wb: Workbook, products: list[Any]) -> None:
    """Sheet «Панели» — все панели всех изделий."""
    ws = wb.create_sheet("Панели")
    ws.append(["ID изделия", "Панель", "Ширина, мм", "Высота, мм", "Толщина, мм", "Материал", "Кромка, мм", "Заметки"])
    for p in products:
        for panel in getattr(p, "panels", []):
            ws.append([
                str(p.id),
                getattr(panel, "name", "") or "",
                getattr(panel, "width_mm", 0),
                getattr(panel, "height_mm", 0),
                getattr(panel, "thickness_mm", 0),
                getattr(panel, "material", "") or "",
                getattr(panel, "edge_band_mm", None) or "",
                getattr(panel, "notes", "") or "",
            ])
    _style_header(ws)


def _write_hardware_sheet(wb: Workbook, bom_items: list[Any]) -> None:
    """Sheet «Фурнитура» — BOM-позиции."""
    ws = wb.create_sheet("Фурнитура")
    ws.append(["SKU", "Наименование", "Кол-во", "Ед. изм.", "Артикул поставщика"])
    for item in bom_items:
        ws.append([
            getattr(item, "sku", "") or "",
            getattr(item, "name", "") or "",
            getattr(item, "qty", 0),
            getattr(item, "unit", "") or "",
            getattr(item, "supplier_sku", "") or "",
        ])
    _style_header(ws)


def generate_order_excel(
    order: Any,
    products: list[Any],
    bom_items: list[Any],
    *,
    is_preview: bool = False,
) -> bytes:
    """Сформировать Excel (.xlsx) для 1C.

    Листы: Заказ, Изделия, Панели, Фурнитура.
    is_preview=True добавляет водяной знак "Предупреждение" (для preview drafts).
    """
    wb = Workbook()
    # убираем дефолтный пустой лист
    wb.remove(wb.active)

    if is_preview:
        # Watermark sheet for preview drafts (Task 12 requirement)
        ws_warn = wb.create_sheet("Предупреждение", 0)
        ws_warn["A1"] = "ВОДЯНОЙ ЗНАК: ПРЕДВАРИТЕЛЬНЫЙ ЧЕРНОВИК — НЕ ДЛЯ ПРОИЗВОДСТВА"
        ws_warn["A2"] = "Экспорт выполнен из черновой ревизии спецификации."
        ws_warn["A3"] = "Для производства используйте только утверждённую ревизию."
        ws_warn["A1"].font = Font(bold=True, color="C00000", size=12)
        ws_warn["A2"].font = Font(italic=True)
        # style header not needed for warn

    _write_order_sheet(wb, order)
    _write_products_sheet(wb, products)
    _write_panels_sheet(wb, products)
    _write_hardware_sheet(wb, bom_items)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── CSV multi-file export ───────────────────────────────────────────

_CSV_PARAMS = dict(delimiter=";", quoting=csv.QUOTE_MINIMAL)


def generate_order_csv(
    order: Any,
    products: list[Any],
    bom_items: list[Any],
    *,
    is_preview: bool = False,
) -> dict[str, str]:
    """Сформировать набор CSV-файлов для 1C.

    Возвращает dict вида {имя_файла: csv_content_str}.
    Файлы: Заказ.csv, Изделия.csv, Панели.csv, Фурнитура.csv.
    is_preview добавляет ПРЕДУПРЕЖДЕНИЕ.txt с watermark (для preview drafts).
    """
    result: dict[str, str] = {}

    if is_preview:
        result["ПРЕДУПРЕЖДЕНИЕ.txt"] = (
            "\ufeffВОДЯНОЙ ЗНАК;ПРЕДВАРИТЕЛЬНЫЙ ЧЕРНОВИК — НЕ ДЛЯ ПРОИЗВОДСТВА\n"
            "Экспорт из draft. Утвердите ревизию перед production use.\n"
        )

    # ── Заказ.csv ──
    buf = io.StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf, **_CSV_PARAMS)
    w.writerow(["Поле", "Значение"])
    w.writerow(["ID заказа", str(order.id)])
    w.writerow(["Клиент", getattr(order, "customer_ref", "") or ""])
    w.writerow(["Статус", getattr(order, "status", "")])
    w.writerow(["Создан", str(order.created_at) if getattr(order, "created_at", None) else ""])
    w.writerow(["Кол-во изделий", len(products)])
    result["Заказ.csv"] = buf.getvalue()

    # ── Изделия.csv ──
    buf = io.StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf, **_CSV_PARAMS)
    w.writerow(["ID", "Название", "Ширина, мм", "Высота, мм", "Глубина, мм", "Материал", "Толщина, мм", "Заметки"])
    for p in products:
        w.writerow([
            str(p.id),
            getattr(p, "name", "") or "",
            getattr(p, "width_mm", 0),
            getattr(p, "height_mm", 0),
            getattr(p, "depth_mm", 0),
            getattr(p, "material", "") or "",
            getattr(p, "thickness_mm", None) or "",
            getattr(p, "notes", "") or "",
        ])
    result["Изделия.csv"] = buf.getvalue()

    # ── Панели.csv ──
    buf = io.StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf, **_CSV_PARAMS)
    w.writerow(["ID изделия", "Панель", "Ширина, мм", "Высота, мм", "Толщина, мм", "Материал", "Кромка, мм", "Заметки"])
    for p in products:
        for panel in getattr(p, "panels", []):
            w.writerow([
                str(p.id),
                getattr(panel, "name", "") or "",
                getattr(panel, "width_mm", 0),
                getattr(panel, "height_mm", 0),
                getattr(panel, "thickness_mm", 0),
                getattr(panel, "material", "") or "",
                getattr(panel, "edge_band_mm", None) or "",
                getattr(panel, "notes", "") or "",
            ])
    result["Панели.csv"] = buf.getvalue()

    # ── Фурнитура.csv ──
    buf = io.StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf, **_CSV_PARAMS)
    w.writerow(["SKU", "Наименование", "Кол-во", "Ед. изм.", "Артикул поставщика"])
    for item in bom_items:
        w.writerow([
            getattr(item, "sku", "") or "",
            getattr(item, "name", "") or "",
            getattr(item, "qty", 0),
            getattr(item, "unit", "") or "",
            getattr(item, "supplier_sku", "") or "",
        ])
    result["Фурнитура.csv"] = buf.getvalue()

    return result


# ── Production gate + CAM identifier contract (Task 12) ─────────────────


PRODUCTION_BLOCK_MSG = (
    "Production export заблокирован для stale draft. "
    "Требуется утверждённая текущая ревизия. "
    "Preview drafts получают watermark."
)


def assert_production_export_allowed(is_approved: bool, context: str = "1c") -> None:
    """Блок production export для stale drafts.

    Используется для 1C и edge экспортов из одной спецификации.
    """
    if not is_approved:
        raise PermissionError(f"[{context}] {PRODUCTION_BLOCK_MSG}")


# Contract identifiers (BOM / 1C / edge / CAM manifest must agree on these)
# No CAM/G-code implemented here.
CAM_MANIFEST_CONTRACT = {
    "panel_id": "stable from ManufacturingSpec.panels[].id",
    "record_id": "revision-aware sha id (e.g. 1c_xxx_rN_ or edge_xxx_rN_)",
    "units": "mm",
    "quantity": 1,
    "revision": "int, part of stable id and cross-check",
    "decimal_sep_options": [".", ","],
    "encoding": "utf-8-sig (BOM)",
    "delimiter": ";",
}


def get_cam_identifier_contract() -> dict:
    """Stable export identifiers contract (for BOM/Excel/CSV/CAM alignment)."""
    return CAM_MANIFEST_CONTRACT
