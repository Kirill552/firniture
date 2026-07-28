from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .models import Supplier, SupplierPrice

_SKU_HEADERS = ("артикул", "код", "sku", "article", "арт")
_PRICE_HEADERS = ("цена", "розница", "опт", "price", "retail", "wholesale")
_CURRENCY_HEADERS = ("валюта", "currency")
_UNIT_HEADERS = ("единица", "ед изм", "ед. изм.", "unit")
_DATE_HEADERS = ("дата прайса", "дата", "price date", "date")


@dataclass(frozen=True)
class PriceRow:
    sku: str
    price: Decimal
    currency: str
    unit: str
    price_date: date
    source_filename: str


@dataclass(frozen=True)
class RejectedPriceRow:
    row_number: int
    reason: str
    values: dict[str, Any]


@dataclass(frozen=True)
class PriceImportResult:
    rows: list[PriceRow]
    rejected: list[RejectedPriceRow]

    @property
    def imported(self) -> int:
        return len(self.rows)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", " ", text).strip()


def _find_column(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    normalized = [_normalize_header(header) for header in headers]
    for alias in aliases:
        target = _normalize_header(alias)
        for index, header in enumerate(normalized):
            if header == target or target in header:
                return index
    return None


def _parse_decimal(value: Any) -> Decimal:
    text = str(value or "").strip().replace(" ", "").replace("\u00a0", "")
    text = text.replace(",", ".")
    return Decimal(text)


def _parse_date(value: Any, fallback: date) -> date:
    if value in (None, ""):
        return fallback
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("не распознана дата прайса")


def _read_rows(path: Path) -> list[list[Any]]:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        return [list(row) for row in next(iter(workbook.worksheets)).iter_rows(values_only=True)]
    if path.suffix.lower() != ".csv":
        raise ValueError("поддерживаются только CSV и XLSX")
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        sample = stream.read(4096)
        stream.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
        except csv.Error:
            dialect = csv.excel
        return [row for row in csv.reader(stream, dialect)]


def import_price_file(path: str | Path, supplier_name: str, price_date: date | None = None) -> PriceImportResult:
    """Разбирает прайс без записи в БД; сохранение выполняет save_price_import."""
    del supplier_name  # имя используется при сохранении, парсеру оно не нужно
    source = Path(path)
    rows = _read_rows(source)
    if not rows:
        return PriceImportResult([], [])
    headers = rows[0]
    sku_column = _find_column(headers, _SKU_HEADERS)
    price_column = _find_column(headers, _PRICE_HEADERS)
    currency_column = _find_column(headers, _CURRENCY_HEADERS)
    unit_column = _find_column(headers, _UNIT_HEADERS)
    date_column = _find_column(headers, _DATE_HEADERS)
    effective_date = price_date or date.today()
    accepted: list[PriceRow] = []
    rejected: list[RejectedPriceRow] = []
    for row_number, values in enumerate(rows[1:], start=2):
        data = {str(headers[index] or ""): values[index] if index < len(values) else "" for index in range(len(headers))}
        if sku_column is None or not str(values[sku_column] if sku_column < len(values) else "").strip():
            rejected.append(RejectedPriceRow(row_number, "не указан артикул", data))
            continue
        if price_column is None:
            rejected.append(RejectedPriceRow(row_number, "не найдена колонка цены", data))
            continue
        try:
            sku = str(values[sku_column]).strip()
            price = _parse_decimal(values[price_column] if price_column < len(values) else "")
            parsed_date = _parse_date(values[date_column] if date_column is not None and date_column < len(values) else "", effective_date)
        except (InvalidOperation, ValueError):
            rejected.append(RejectedPriceRow(row_number, "не распознана цена или дата", data))
            continue
        accepted.append(PriceRow(sku, price, str(values[currency_column]).strip() if currency_column is not None and currency_column < len(values) else "RUB", str(values[unit_column]).strip() if unit_column is not None and unit_column < len(values) else "шт", parsed_date, source.name))
    return PriceImportResult(accepted, rejected)


async def save_price_import(session: AsyncSession, supplier_name: str, result: PriceImportResult) -> int:
    supplier = await session.scalar(select(Supplier).where(Supplier.name == supplier_name))
    if supplier is None:
        supplier = Supplier(name=supplier_name)
        session.add(supplier)
        await session.flush()
    session.add_all([SupplierPrice(sku=row.sku, supplier_id=supplier.id, price=row.price, currency=row.currency, unit=row.unit, price_date=row.price_date, source_filename=row.source_filename) for row in result.rows])
    await session.commit()
    return len(result.rows)


async def get_current_price(session: AsyncSession, sku: str, supplier_name: str | None = None) -> Decimal | None:
    query = select(SupplierPrice.price).join(Supplier).where(SupplierPrice.sku == sku)
    if supplier_name:
        query = query.where(Supplier.name == supplier_name)
    query = query.order_by(SupplierPrice.price_date.desc(), SupplierPrice.created_at.desc()).limit(1)
    return await session.scalar(query)
