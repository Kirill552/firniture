"""Tests for 1C export core — Task 12.

Focused tests for stable IDs, Russian CSV output, and export decision domain.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import UTC, datetime
from uuid import uuid4

import pytest

from api.export_1c import (
    Export1CRecord,
    ExportStatus,
    assert_production_export_allowed,
    build_1c_records,
    export_1c_csv,
    export_1c_csv_bytes,
    generate_1c_filename,
    generate_order_csv,
    generate_order_excel,
    get_cam_identifier_contract,
    make_1c_record_id,
)
from api.manufacturing.contracts import ManufacturingSpec, PanelSpec

# ── Fixtures ──────────────────────────────────────────────────────────

SIMPLE_SPEC = ManufacturingSpec(
    spec_version="1.0",
    panels=[
        PanelSpec(id="panel_1", width_mm=600, height_mm=720, thickness_mm=16, material="ЛДСП"),
        PanelSpec(id="panel_2", width_mm=300, height_mm=720, thickness_mm=16, material="ЛДСП"),
    ],
)

EMPTY_SPEC = ManufacturingSpec(spec_version="1.0", panels=[])

SINGLE_PANEL_SPEC = ManufacturingSpec(
    spec_version="1.0",
    panels=[
        PanelSpec(id="side_left", width_mm=500, height_mm=800, thickness_mm=18, material="МДФ"),
    ],
)

NO_MATERIAL_SPEC = ManufacturingSpec(
    spec_version="1.0",
    panels=[
        PanelSpec(id="bare_panel", width_mm=400, height_mm=600, thickness_mm=16),
    ],
)


# ── ExportStatus enum ─────────────────────────────────────────────────


class TestExportStatus1C:
    def test_values(self):
        assert ExportStatus.DRAFT.value == "draft"
        assert ExportStatus.APPROVED.value == "approved"

    def test_from_string(self):
        assert ExportStatus("draft") is ExportStatus.DRAFT
        assert ExportStatus("approved") is ExportStatus.APPROVED

    def test_invalid_raises(self):
        with pytest.raises(ValueError):
            ExportStatus("pending")


# ── Stable IDs ────────────────────────────────────────────────────────


class TestStableIds1C:
    def test_deterministic(self):
        id1 = make_1c_record_id("panel_1")
        id2 = make_1c_record_id("panel_1")
        assert id1 == id2

    def test_different_panels_different_ids(self):
        id1 = make_1c_record_id("panel_1")
        id2 = make_1c_record_id("panel_2")
        assert id1 != id2

    def test_format(self):
        rid = make_1c_record_id("panel_1")
        # revision-aware (default rev=1)
        assert rid.startswith("1c_panel_1_r1_")
        assert len(rid.split("_")[-1]) == 12


# ── Export1CRecord validation ─────────────────────────────────────────


class TestExport1CRecord:
    def test_defaults(self):
        rec = Export1CRecord(
            record_id="r1",
            panel_id="p1",
            name="Test",
            width_mm=600,
            height_mm=720,
            thickness_mm=16,
            material="ЛДСП",
        )
        assert rec.quantity == 1
        assert rec.status is ExportStatus.DRAFT

    def test_empty_record_id_raises(self):
        with pytest.raises(ValueError):
            Export1CRecord(
                record_id="",
                panel_id="p1",
                name="Test",
                width_mm=600,
                height_mm=720,
                thickness_mm=16,
                material="ЛДСП",
            )

    def test_zero_quantity_raises(self):
        with pytest.raises(ValueError):
            Export1CRecord(
                record_id="r1",
                panel_id="p1",
                name="Test",
                width_mm=600,
                height_mm=720,
                thickness_mm=16,
                material="ЛДСП",
                quantity=0,
            )

    def test_negative_quantity_raises(self):
        with pytest.raises(ValueError):
            Export1CRecord(
                record_id="r1",
                panel_id="p1",
                name="Test",
                width_mm=600,
                height_mm=720,
                thickness_mm=16,
                material="ЛДСП",
                quantity=-1,
            )


# ── build_1c_records ──────────────────────────────────────────────────


class TestBuild1CRecords:
    def test_empty_spec(self):
        records = build_1c_records(EMPTY_SPEC)
        assert records == []

    def test_single_panel(self):
        records = build_1c_records(SINGLE_PANEL_SPEC)
        assert len(records) == 1
        rec = records[0]
        assert rec.panel_id == "side_left"
        assert rec.width_mm == 500
        assert rec.height_mm == 800
        assert rec.thickness_mm == 18
        assert rec.material == "МДФ"

    def test_two_panels(self):
        records = build_1c_records(SIMPLE_SPEC)
        assert len(records) == 2

    def test_deterministic_order(self):
        r1 = build_1c_records(SIMPLE_SPEC)
        r2 = build_1c_records(SIMPLE_SPEC)
        ids1 = [r.record_id for r in r1]
        ids2 = [r.record_id for r in r2]
        assert ids1 == ids2

    def test_sorted_by_panel_id(self):
        records = build_1c_records(SIMPLE_SPEC)
        panel_ids = [r.panel_id for r in records]
        assert panel_ids == sorted(panel_ids)

    def test_material_default_ldsp(self):
        records = build_1c_records(NO_MATERIAL_SPEC)
        assert records[0].material == "ЛДСП"

    def test_name_format(self):
        records = build_1c_records(SIMPLE_SPEC)
        assert records[0].name == "Панель panel_1"
        assert records[1].name == "Панель panel_2"

    def test_all_have_stable_ids(self):
        records = build_1c_records(SIMPLE_SPEC)
        for r in records:
            assert r.record_id.startswith("1c_")


# ── CSV export ────────────────────────────────────────────────────────


class TestExport1cCsv:
    def test_starts_with_bom(self):
        csv_str = export_1c_csv(build_1c_records(SIMPLE_SPEC))
        assert csv_str.startswith("\ufeff")

    def test_semicolon_delimiter(self):
        csv_str = export_1c_csv(build_1c_records(SINGLE_PANEL_SPEC))
        lines = csv_str.strip().split("\n")
        assert ";" in lines[1]

    def test_header_row(self):
        csv_str = export_1c_csv(build_1c_records(SINGLE_PANEL_SPEC))
        lines = csv_str.strip().split("\n")
        header = lines[0][1:]  # remove BOM
        assert "Наименование" in header
        assert "Длина, мм" in header
        assert "Ширина, мм" in header
        assert "Толщина, мм" in header
        assert "Материал" in header
        assert "Кол-во" in header
        assert "Статус" in header

    def test_draft_status_by_default(self):
        records = build_1c_records(SIMPLE_SPEC)
        csv_str = export_1c_csv(records)
        assert "Черновик" in csv_str

    def test_approved_status(self):
        records = build_1c_records(SIMPLE_SPEC)
        for r in records:
            r.status = ExportStatus.APPROVED
        csv_str = export_1c_csv(records)
        assert "Утверждено" in csv_str
        assert "Черновик" not in csv_str

    def test_parseable_csv(self):
        records = build_1c_records(SIMPLE_SPEC)
        csv_str = export_1c_csv(records)
        reader = csv.reader(io.StringIO(csv_str[1:]), delimiter=";")
        rows = list(reader)
        # Header + 2 data rows
        assert len(rows) == 3

    def test_empty_records(self):
        csv_str = export_1c_csv([])
        lines = csv_str.strip().split("\n")
        assert len(lines) == 1  # header only

    def test_bytes_output(self):
        records = build_1c_records(SINGLE_PANEL_SPEC)
        csv_bytes = export_1c_csv_bytes(records)
        assert isinstance(csv_bytes, bytes)
        decoded = csv_bytes.decode("utf-8")
        assert decoded.startswith("\ufeff")

    def test_dimensions_in_output(self):
        records = build_1c_records(SINGLE_PANEL_SPEC)
        csv_str = export_1c_csv(records)
        assert "500.0" in csv_str  # width
        assert "800.0" in csv_str  # height
        assert "18.0" in csv_str  # thickness


# ── Lightweight mocks for route-level tests ──────────────────────────


@dataclass
class _MockPanel:
    name: str = "Panel"
    width_mm: float = 600
    height_mm: float = 720
    thickness_mm: float = 16
    material: str = "ЛДСП"
    edge_band_mm: float | None = None
    notes: str = ""


@dataclass
class _MockProduct:
    id: str = field(default_factory=lambda: str(uuid4()))
    name: str = "Изделие"
    width_mm: float = 800
    height_mm: float = 720
    depth_mm: float = 600
    material: str = "ЛДСП"
    thickness_mm: float = 16
    notes: str = ""
    panels: list[_MockPanel] = field(default_factory=list)


@dataclass
class _MockBomItem:
    sku: str = "SKU-001"
    name: str = "Петля"
    qty: float = 2.0
    unit: str = "шт"
    supplier_sku: str = "SUP-001"


@dataclass
class _MockOrder:
    id: uuid4 = field(default_factory=uuid4)
    customer_ref: str = "Клиент ABC"
    status: str = "draft"
    created_at: datetime = field(default_factory=lambda: datetime(2026, 7, 12, tzinfo=UTC))
    products: list[_MockProduct] = field(default_factory=list)


# ── Route-level import regression ────────────────────────────────────


class TestRouteImportRegression:
    """Verify that the three symbols routers.py imports actually exist."""

    def test_generate_1c_filename_importable(self):
        from api.export_1c import generate_1c_filename as fn
        assert callable(fn)

    def test_generate_order_csv_importable(self):
        from api.export_1c import generate_order_csv as fn
        assert callable(fn)

    def test_generate_order_excel_importable(self):
        from api.export_1c import generate_order_excel as fn
        assert callable(fn)


# ── generate_1c_filename ─────────────────────────────────────────────


class TestGenerate1cFilename:
    def test_with_customer_ref(self):
        order = _MockOrder(customer_ref="ООО Ромашка")
        name = generate_1c_filename(order, "xlsx")
        assert name == "1C_Order_ООО_Ромашка_20260712.xlsx"

    def test_without_customer_ref(self):
        order = _MockOrder(customer_ref="")
        name = generate_1c_filename(order, "zip")
        assert name.startswith("1C_Order_")
        assert name.endswith("_20260712.zip")

    def test_special_chars_sanitized(self):
        order = _MockOrder(customer_ref="ООО/Ромашка & Sons")
        name = generate_1c_filename(order, "csv")
        assert "/" not in name
        assert "&" not in name


# ── generate_order_excel ─────────────────────────────────────────────


class TestGenerateOrderExcel:
    def test_returns_bytes(self):
        order = _MockOrder()
        result = generate_order_excel(order, [], [])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self):
        from openpyxl import load_workbook
        order = _MockOrder()
        product = _MockProduct(panels=[_MockPanel()])
        bom = [_MockBomItem()]
        xlsx_bytes = generate_order_excel(order, [product], bom)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_names = wb.sheetnames
        assert "Заказ" in sheet_names
        assert "Изделия" in sheet_names
        assert "Панели" in sheet_names
        assert "Фурнитура" in sheet_names

    def test_order_sheet_has_id(self):
        from openpyxl import load_workbook
        order = _MockOrder()
        xlsx_bytes = generate_order_excel(order, [], [])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Заказ"]
        # Row 2 should contain the order ID
        assert str(order.id) in str(ws.cell(2, 2).value)

    def test_products_sheet_has_data(self):
        from openpyxl import load_workbook
        order = _MockOrder()
        product = _MockProduct(name="Шкаф", width_mm=1200)
        xlsx_bytes = generate_order_excel(order, [product], [])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Изделия"]
        # Row 2: first data row
        assert ws.cell(2, 2).value == "Шкаф"
        assert ws.cell(2, 3).value == 1200

    def test_panels_sheet_has_data(self):
        from openpyxl import load_workbook
        order = _MockOrder()
        product = _MockProduct(panels=[_MockPanel(name="Боковина", width_mm=500)])
        xlsx_bytes = generate_order_excel(order, [product], [])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Панели"]
        assert ws.cell(2, 2).value == "Боковина"
        assert ws.cell(2, 3).value == 500

    def test_hardware_sheet_has_data(self):
        from openpyxl import load_workbook
        order = _MockOrder()
        bom = [_MockBomItem(sku="HINGE-01", name="Петля 110°")]
        xlsx_bytes = generate_order_excel(order, [], bom)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Фурнитура"]
        assert ws.cell(2, 1).value == "HINGE-01"
        assert ws.cell(2, 2).value == "Петля 110°"

    def test_empty_inputs(self):
        order = _MockOrder()
        result = generate_order_excel(order, [], [])
        assert isinstance(result, bytes)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(result))
        assert "Заказ" in wb.sheetnames


# ── generate_order_csv ───────────────────────────────────────────────


class TestGenerateOrderCsv:
    def test_returns_dict_with_expected_keys(self):
        order = _MockOrder()
        result = generate_order_csv(order, [], [])
        assert isinstance(result, dict)
        assert "Заказ.csv" in result
        assert "Изделия.csv" in result
        assert "Панели.csv" in result
        assert "Фурнитура.csv" in result

    def test_all_values_are_strings(self):
        order = _MockOrder()
        result = generate_order_csv(order, [], [])
        for v in result.values():
            assert isinstance(v, str)

    def test_all_have_bom(self):
        order = _MockOrder()
        result = generate_order_csv(order, [], [])
        for content in result.values():
            assert content.startswith("\ufeff")

    def test_order_csv_has_customer_ref(self):
        order = _MockOrder(customer_ref="ТестКлиент")
        result = generate_order_csv(order, [], [])
        assert "ТестКлиент" in result["Заказ.csv"]

    def test_products_csv_has_data(self):
        order = _MockOrder()
        product = _MockProduct(name="Тумба", width_mm=400)
        result = generate_order_csv(order, [product], [])
        assert "Тумба" in result["Изделия.csv"]
        assert "400" in result["Изделия.csv"]

    def test_panels_csv_has_data(self):
        order = _MockOrder()
        product = _MockProduct(panels=[_MockPanel(name="Дверца", width_mm=350)])
        result = generate_order_csv(order, [product], [])
        assert "Дверца" in result["Панели.csv"]
        assert "350" in result["Панели.csv"]

    def test_hardware_csv_has_data(self):
        order = _MockOrder()
        bom = [_MockBomItem(sku="CAM-LOCK", name="Эксцентрик")]
        result = generate_order_csv(order, [], bom)
        assert "CAM-LOCK" in result["Фурнитура.csv"]
        assert "Эксцентрик" in result["Фурнитура.csv"]

    def test_semicolon_delimited(self):
        order = _MockOrder()
        result = generate_order_csv(order, [], [])
        for content in result.values():
            lines = content.strip().split("\n")
            assert ";" in lines[0], f"Expected semicolon delimiter in: {lines[0]}"

    def test_empty_inputs(self):
        order = _MockOrder()
        result = generate_order_csv(order, [], [])
        # Should still produce all 4 files (with headers only)
        assert len(result) == 4


# ── Task 12 additions: revision-aware stable ids, Russian fixtures, block+watermark ──


class Test1CStableIdsWithRevisionAndCAM:
    def test_make_id_with_revision(self):
        rid = make_1c_record_id("side", revision=3)
        assert "r3" in rid
        assert rid.startswith("1c_side_r3_")

    def test_build_records_with_revision_and_status(self):
        recs = build_1c_records(SIMPLE_SPEC, revision=9, status=ExportStatus.APPROVED)
        assert len(recs) == 2
        assert all("r9" in r.record_id for r in recs)
        assert recs[0].status is ExportStatus.APPROVED

    def test_cam_contract_for_identifiers(self):
        c = get_cam_identifier_contract()
        assert "revision" in c
        assert c["units"] == "mm"
        assert c["delimiter"] == ";"


class TestRussianFixturesDecimalAnd1CExpectations:
    """Decimal separators, encoding, column names, 1C expectations (RU fixtures)."""

    def test_1c_csv_decimal_comma(self):
        recs = build_1c_records(SINGLE_PANEL_SPEC)
        csv_str = export_1c_csv(recs, decimal_sep=",")
        assert "500,0" in csv_str
        assert "800,0" in csv_str
        assert "18,0" in csv_str

    def test_1c_csv_bom_encoding_and_columns(self):
        csv_str = export_1c_csv(build_1c_records(SINGLE_PANEL_SPEC))
        assert csv_str.startswith("\ufeff")
        # Russian column names for 1C
        assert "ИД записи" in csv_str
        assert "Длина, мм" in csv_str
        assert "Кол-во" in csv_str
        assert "Статус" in csv_str

    def test_1c_semicolon_compatible(self):
        csv = export_1c_csv(build_1c_records(SIMPLE_SPEC))
        data_line = csv.split("\n")[1]
        assert data_line.count(";") >= 7


class Test1CProductionBlockAndPreviewWatermark:
    def test_block_production_stale_draft(self):
        with pytest.raises(PermissionError) as ei:
            assert_production_export_allowed(is_approved=False)
        msg = str(ei.value)
        assert "Production export" in msg or "заблокирован" in msg

    def test_allow_production_when_approved(self):
        assert_production_export_allowed(True)  # no raise

    def test_watermark_in_1c_preview_csv(self):
        csv = export_1c_csv(build_1c_records(SIMPLE_SPEC), watermark=True)
        assert "ВОДЯНОЙ_ЗНАК" in csv
        assert "ПРЕДВАРИТЕЛЬНЫЙ ЧЕРНОВИК" in csv

    def test_preview_watermark_excel_adds_sheet(self):
        order = _MockOrder()
        xlsx = generate_order_excel(order, [], [], is_preview=True)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        assert "Предупреждение" in wb.sheetnames
        ws = wb["Предупреждение"]
        assert "ВОДЯНОЙ ЗНАК" in str(ws["A1"].value)
        assert "НЕ ДЛЯ ПРОИЗВОДСТВА" in str(ws["A1"].value)

    def test_preview_watermark_in_order_csv_bundle(self):
        order = _MockOrder()
        res = generate_order_csv(order, [], [], is_preview=True)
        assert "ПРЕДУПРЕЖДЕНИЕ.txt" in res
        assert "ВОДЯНОЙ ЗНАК" in res["ПРЕДУПРЕЖДЕНИЕ.txt"]
        # still have the 4 data files
        assert "Заказ.csv" in res
